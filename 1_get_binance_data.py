import numpy as np
import pandas as pd
from binance.client import Client
from datetime import datetime as dt
from datetime import timedelta

api_key = "XTHaOWGOcfEYmEnNpcVPJ7LsQgSnk6eBKo4UiOB62w0HfxFqPqkXimPK79lS8eHf"
api_secret = "0jlajqe8UuKIEd5Rew6w2SpimZEaJ9OEAHlaAxneBkGboEOuktTpecXv6XRiIGYL"

client = Client(api_key, api_secret)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, index=False, header=None, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()



def scrape_hourly_candles(market="BTCUSDT", startingDate = dt(2018,9,14)):
    date = startingDate  
    
    dateList = []
    while date<dt.today():
        date += timedelta(days=1)  
        dateList.append(dt.strftime(date, "%Y-%m-%d"))
    
    i = 0
    row = 0
    
    while i<len(dateList)-1: 
         #klines = client.get_historical_klines(market, client.KLINE_INTERVAL_1HOUR, start_str = dateList[i], end_str=dateList[i+1])
         klines = client.get_historical_klines(market, '1m', start_str = dateList[i], end_str=dateList[i+1])
         try:
             klines.pop(0)
         except:
             pass
         k=np.array(klines)
         time = ''
         times = []
         for j in range(0,24):
             for jj in range(0,60):
                 if (j<10 and jj<10):
                     time = '0' + str(j) + ':' + '0' + str(jj)
                 elif (j<10):
                     time = '0' + str(j) + ':' + str(jj)
                 elif (jj<10):
                     time = str(j) + ':' + '0' + str(jj)
                 else:
                     time = str(j) + ':' + str(jj)

                 times.append(dateList[i]+" "+time)

         if len(k)==1440:
             df = pd.DataFrame({"date": times, 
                                "volume": k[:,0],
                                "open": k[:,1],
                                "high": k[:,2],
                                "low": k[:,3],
                                "close": k[:,4]
                                })
             
             #append_df_to_excel(market+".xlsx", df, startrow=row)
             df.to_csv(market+".csv", mode='a', header=False)
             row+=1440
         i+=1
         print(len(k), " saved "+str(row)+" rows of data ", dateList[i], " ", dateList[i+1])

def scrape_minute_candles(market="BTCUSDT", startingDate = dt(2018,9,14)):
    date = startingDate  
    
    dateList = []
    while date<dt.today():
        date += timedelta(days=1)  
        dateList.append(dt.strftime(date, "%Y-%m-%d"))
    
    i = 0
    row = 0
    
    while i<len(dateList)-1: 
         #klines = client.get_historical_klines(market, client.KLINE_INTERVAL_1HOUR, start_str = dateList[i], end_str=dateList[i+1])
         klines = client.get_historical_klines(market, '1m', start_str = dateList[i], end_str=dateList[i+1])
         try:
             klines.pop(0)
         except:
             pass
         k=np.array(klines)
         time = ''
         times = []
         for j in range(0,24):
             for jj in range(0,60):
                 if (j<10 and jj<10):
                     time = '0' + str(j) + ':' + '0' + str(jj)
                 elif (j<10):
                     time = '0' + str(j) + ':' + str(jj)
                 elif (jj<10):
                     time = str(j) + ':' + '0' + str(jj)
                 else:
                     time = str(j) + ':' + str(jj)

                 times.append(dateList[i]+" "+time)

         if len(k)==1440:
             df = pd.DataFrame({"date": times, 
                                "volume": k[:,0],
                                "open": k[:,1],
                                "high": k[:,2],
                                "low": k[:,3],
                                "close": k[:,4]
                                })
             
             #append_df_to_excel(market+".xlsx", df, startrow=row)
             df.to_csv(market+".csv", mode='a', header=False)
             row+=1440
         i+=1
         print(len(k), " saved "+str(row)+" rows of data ", dateList[i], " ", dateList[i+1])

#scrape_hourly_candles(market="SHIBEUR", startingDate=dt(2021,5,1)) # max 2 years of data
scrape_minute_candles(market="SHIBEUR", startingDate=dt(2021,5,1)) # max 60 days of data
